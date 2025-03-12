import pandas as pd
import math
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

class ExcelSplitterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Splitter")
        self.root.geometry("650x450")
        self.root.resizable(True, True)
        
        # Variabel untuk menyimpan path file
        self.input_file_path = tk.StringVar()
        self.output_directory = tk.StringVar()
        self.rows_per_file = tk.StringVar(value="500")
        
        # Setup UI
        self.setup_styles()
        self.create_widgets()
        
        # Inisialisasi log
        self.log("Aplikasi siap digunakan. Silakan pilih file Excel untuk di-split.")

    def setup_styles(self):
        style = ttk.Style()
        
        # Warna tema
        primary_color = "#4a86e8"
        
        # Konfigurasi tombol utama
        style.configure("Accent.TButton", 
                       font=("Arial", 11, "bold"),
                       padding=(20, 10),
                       background=primary_color)
        
        # Tombol aksi
        style.configure("Action.TButton",
                       padding=(10, 5))
        
        # Label header
        style.configure("Header.TLabel", 
                       foreground=primary_color)
        
        # Label frame header
        style.configure("TLabelframe.Label", 
                       font=("Arial", 10, "bold"),
                       foreground=primary_color)

    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill="x", pady=(0, 15))
        
        ttk.Label(header_frame, text="Excel File Splitter", 
                 font=("Arial", 18, "bold"), style="Header.TLabel").pack(pady=(0, 5))
        ttk.Label(header_frame, text="Split Excel files into smaller chunks", 
                 font=("Arial", 10)).pack()
        ttk.Label(header_frame, text="(Only take Reference and Tax Invoice Number)", 
                 font=("Arial", 9, "italic"), foreground="#666666").pack()
        ttk.Label(header_frame, text="by Mahendra", 
                 font=("Arial", 9), foreground="#666666").pack(pady=(5,0))
        
        # Separator
        ttk.Separator(header_frame, orient="horizontal").pack(fill="x", pady=(10, 0))
        
        # Input frame
        input_frame = ttk.LabelFrame(main_frame, text="File Settings", padding=(15, 10))
        input_frame.pack(fill="x", pady=10)
        
        # Input file
        file_frame = ttk.Frame(input_frame)
        file_frame.pack(fill="x", pady=(5, 10))
        ttk.Label(file_frame, text="Input Excel:", width=12).pack(side="left", padx=5)
        ttk.Entry(file_frame, textvariable=self.input_file_path, width=50).pack(side="left", padx=5, fill="x", expand=True)
        ttk.Button(file_frame, text="Browse", command=self.browse_input_file, style="Action.TButton").pack(side="left", padx=(5, 0))
        
        # Output directory
        output_frame = ttk.Frame(input_frame)
        output_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(output_frame, text="Output Folder:", width=12).pack(side="left", padx=5)
        ttk.Entry(output_frame, textvariable=self.output_directory, width=50).pack(side="left", padx=5, fill="x", expand=True)
        ttk.Button(output_frame, text="Browse", command=self.browse_output_dir, style="Action.TButton").pack(side="left", padx=(5, 0))
        
        # Rows per file
        rows_frame = ttk.Frame(input_frame)
        rows_frame.pack(fill="x")
        ttk.Label(rows_frame, text="Rows per file:", width=12).pack(side="left", padx=5)
        ttk.Entry(rows_frame, textvariable=self.rows_per_file, width=10).pack(side="left", padx=5)
        
        # Split button
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        split_button = ttk.Button(button_frame, text="Split File", command=self.split_file, style="Accent.TButton")
        split_button.pack(pady=5, padx=100)
        
        # Log area
        log_frame = ttk.LabelFrame(main_frame, text="Log", padding=(10, 5))
        log_frame.pack(fill="both", expand=True, pady=10)
        
        # Scrollbar dan Text widget untuk log
        self.log_text = tk.Text(log_frame, height=8, wrap="word", 
                               font=("Consolas", 9), bg="#f9f9f9", fg="#333333")
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(side="left", fill="both", expand=True)

    def browse_input_file(self):
        filename = filedialog.askopenfilename(
            title="Pilih File Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.input_file_path.set(filename)
            # Set default output directory ke folder yang sama
            self.output_directory.set(os.path.dirname(filename))
            self.log(f"File input dipilih: {filename}")

    def browse_output_dir(self):
        directory = filedialog.askdirectory(title="Pilih Folder Output")
        if directory:
            self.output_directory.set(directory)
            self.log(f"Folder output dipilih: {directory}")

    def log(self, message):
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)

    def split_file(self):
        try:
            input_file = self.input_file_path.get()
            output_dir = self.output_directory.get()
            rows_per_file = int(self.rows_per_file.get())

            if not input_file or not output_dir:
                messagebox.showerror("Error", "Pilih file input dan folder output terlebih dahulu!")
                return

            self.log("Membaca file Excel...")
            df = pd.read_excel(input_file, dtype=str)

            # Pilih kolom yang dibutuhkan
            cols_to_extract = ['refrence AQ', 'TaxInvoiceNumber']
            if set(cols_to_extract).issubset(df.columns):
                df_selected = df[cols_to_extract].copy()
                df_selected.columns = ['No IV', 'No FP']
            else:
                self.log("Menggunakan indeks kolom alternatif...")
                df_selected = df.iloc[:, [42, 13]].copy()
                df_selected.columns = ['No IV', 'No FP']

            # Filter baris kosong
            df_selected = df_selected.replace('', pd.NA)
            df_selected = df_selected.dropna()

            # Hitung jumlah file
            total_baris = len(df_selected)
            jumlah_file = math.ceil(total_baris / rows_per_file)

            self.log(f"Total {total_baris} baris akan dibagi menjadi {jumlah_file} file")

            # Proses splitting
            for i in range(jumlah_file):
                start_idx = i * rows_per_file
                end_idx = min((i + 1) * rows_per_file, total_baris)
                
                df_chunk = df_selected.iloc[start_idx:end_idx]
                output_file = os.path.join(output_dir, f'Report_NoIV_NoFP_{i+1}.xlsx')
                
                df_chunk.to_excel(output_file, index=False)
                
                # Format Excel
                wb = load_workbook(output_file)
                ws = wb.active
                
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                for cell in ws[1]:
                    cell.style = 'Headline 3'
                
                wb.save(output_file)
                self.log(f'File {i+1} dari {jumlah_file} selesai dibuat')

            messagebox.showinfo("Sukses", f"Berhasil membuat {jumlah_file} file di folder {output_dir}")
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSplitterApp(root)
    root.mainloop()