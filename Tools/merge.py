import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

def merge_excel_files(input_folder, output_file):
    pd.options.display.float_format = '{:.0f}'.format
    
    all_dataframes = []
    
    for file in os.listdir(input_folder):
        if file.endswith('.xlsx'):
            file_path = os.path.join(input_folder, file)
            df = pd.read_excel(file_path, dtype=str)
            df['Source_File'] = file
            all_dataframes.append(df)
    
    if all_dataframes:
        merged_df = pd.concat(all_dataframes, ignore_index=True)
        merged_df.to_excel(output_file, index=False)
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                    
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        for cell in ws[1]:
            cell.style = 'Headline 3'
        
        wb.save(output_file)
        print(f"File berhasil digabungkan dan disimpan ke {output_file}")
    else:
        print("Tidak ada file Excel yang ditemukan di folder")

if __name__ == "__main__":
    input_folder = r"D:\Coretax Chandra\04-9-1-2025"
    output_file = "D:/Coretax Chandra/04-9-1-2025/04-9-1-2025.xlsx"
    
    merge_excel_files(input_folder, output_file)
