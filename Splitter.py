import pandas as pd
import math
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Baca file Excel dengan dtype=str untuk menghindari konversi otomatis
file_path = 'Report detail Listing 1-31  jan 2024 - NOFPUPDATE.xls'
df = pd.read_excel(file_path, dtype=str)

# Dapatkan header
header = df.columns

# Hitung jumlah file yang akan dibuat
baris_per_file = 500
total_baris = len(df)
jumlah_file = math.ceil(total_baris / baris_per_file)

# Pisahkan data menjadi beberapa file
for i in range(jumlah_file):
    # Tentukan index awal dan akhir untuk setiap chunk
    start_idx = i * baris_per_file
    end_idx = min((i + 1) * baris_per_file, total_baris)
    
    # Ambil chunk data
    df_chunk = df.iloc[start_idx:end_idx]
    
    # Buat nama file baru
    output_file = f'Report_detail_{i+1}.xlsx'
    
    # Simpan ke file Excel baru
    df_chunk.to_excel(output_file, index=False)
    
    # Format file Excel yang baru dibuat
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Format kolom dan angka
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                # Format angka dengan pemisah ribuan
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format header
    for cell in ws[1]:
        cell.style = 'Headline 3'
    
    wb.save(output_file)
    print(f'File {output_file} telah dibuat dengan {len(df_chunk)} baris')

print(f'\nTotal {jumlah_file} file telah dibuat')